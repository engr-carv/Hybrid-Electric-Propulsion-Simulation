import numpy as np
from dataclasses import dataclass
import pandas as pd
from openpyxl import load_workbook
import sys
import os

def directory_check():
    """Check if the 'output_data' and 'input_data' directories exist. If not, create them, along with the 'plots' subfolder inside 'output_data'."""
    # Define the directories with full paths
    directories = [os.path.join('..', 'output_data'), os.path.join('..', 'input_data')]
    
    # Check and create the directories if needed
    for directory in directories:
        if not os.path.exists(directory):
            os.makedirs(directory)
            print(f"Directory '{directory}' created.")
        else:
            print(f"Directory '{directory}' already exists.")
    
    # Now, add the 'plots' subfolder under 'output_data'
    plots_directory = os.path.join('..', 'output_data', 'plots')
    if not os.path.exists(plots_directory):
        os.makedirs(plots_directory)
        print("Subfolder 'plots' inside 'output_data' created.")
    else:
        print("Subfolder 'plots' inside 'output_data' already exists.")

def isa_properties(altitude, gamma):
    """
    Calculate ISA atmospheric properties at a given altitude.

    Parameters:
    altitude (float): Altitude in feet.

    Returns:
    tuple: Temperature (K), Pressure (Pa), Density (kg/m³), Speed of Sound (m/s)
    """
    # Constants
    T0 = 288.15 # Sea level standard temperature [K]
    P0 = 101325 # Sea level standard pressure [Pa]
    L = 0.0065 # Temperature lapse rate [K/m]
    R = 287.05  # Specific gas constant for dry air [J/kg*K]

    # Calculate temperature and pressure based on altitude
    T = T0 - (L * altitude)  # Altitude is in meters
    P = P0 * (1 - (L * altitude) / T0) ** (9.81/(R * L)) # Altitude is in meters

    # Calculate density
    rho = P / (R * T)

    # Calculate speed of sound
    a = (gamma * R * T) ** 0.5                                                    # Assuming air as a perfect gas

    return rho, a

@dataclass()
class MissionCharacteristics:
    a: np.ndarray  # Speed of sound array
    rho: np.ndarray  # Density array

def mission_char(cruise_alt, gamma):
    """
    Calculate mission characteristics including speed of sound and air density.

    Parameters:
    cruise_alt (float): Cruise altitude in meters.

    Returns:
    MissionCharacteristics: A dataclass containing speed of sound and density.
    """
    # Generate altitude range in meters
    altitude_range = np.arange(0, cruise_alt + 1, 0.1)                               # Altitudes from ground level to cruise

    # Initialize arrays
    speeds_of_sound = np.zeros((len(altitude_range), 1), dtype='d')
    densities = np.zeros((len(altitude_range), 1), dtype='d')
    
    '''
     Calculate atmospheric properties for each altitude
     calculates the speed of sound for only a certain range of altitudes, but
     calculates the density for the entire range of altitudes
    '''
    for i, alt in enumerate(altitude_range):
        rho, a = isa_properties(alt, gamma)
        speeds_of_sound[i] = a
        densities[i] = rho

    return MissionCharacteristics(a=speeds_of_sound, rho=densities)

def prop_char(power_thru, phi, spe, batt_m, eta_therm, eta_g, eta_inv, eta_c, eta_em, 
              eta_prop):
    """
    Create a dictionary to store and organize all characteristics of the 
    propulsion system
    """
    prp = {}

    # Inherent Characteristics of the Battery
    prp['batt'] = {
        'phi': phi,                                                             # Degree of Hybridization [-]
        'minSoCpercentage': 0.2,                                                # Minimum state of charge for battery safety
        'spe': spe,                                                             # Convert Wh/kg to J/kg
        'm': batt_m,                                                            # Battery mass [kg]
        'power_thru': power_thru,                                               # Power throughput battery is capable of [W]
    }
    
    full_batt_cap = prp['batt']['spe'] * prp['batt']['m']                       # Raw battery capacity without accounting for min SoC [J]
    prp['full_charge'] = full_batt_cap * (1 - prp['batt']['minSoCpercentage'])  # Usable battery capacity after accounting for min SoC [J]
    prp['empty_charge'] = prp['full_charge'] *  prp['batt']['minSoCpercentage'] # Depleted state of the battery to maximizes logevity and capacity [J]
    
    # Turboshaft Engine
    prp['ts'] = {'delH_f': 43e6}                                                # Heating value of kerosene [J/kg]
    
    prp['eta_batt_to_prop'] = eta_inv * eta_em * eta_prop * eta_c               # The efficiency to transfer energy from the battery to the propulsors
    
    prp['eta_ts_to_prop'] = eta_therm * eta_g * eta_c * eta_em * eta_prop   # The efficiency to transfer energy from the turboshaft engine to the propulsors
    prp['eta_ts_to_charge'] = eta_therm * eta_c * eta_g * eta_inv           # The efficiency to transfer energy from the turboshaft engine to the battery for charging

    return prp

def actuatorDiskTheory(T, rhoAtAlt, trueVelocity, Dprop):
    eta_prop = 2 / \
    (1 + np.sqrt(1 + (T / (0.5 * rhoAtAlt * trueVelocity^2 * (np.pi / 4) * Dprop^2))))
    
    eta_prop = 0.9 * eta_prop
    
    return eta_prop

def charge_batt(P, t, eta_ts_to_prop, delH_f, batt_thru, eta_ts_to_charge, cap):
    ''' Calculates the fuel burn for charging the battery and adds capacity to said battery '''
    m_fuel = ((P * t) + (batt_thru * t)) / \
        ((eta_ts_to_prop + eta_ts_to_charge) * \
         delH_f)                                                                # Fuel burnt is that of both charging the battery and providing propulsive power
    cap += batt_thru * t  
    return cap, m_fuel

def power_logger(powerArray, PowerToMaintainFlight, hybridizationFactor,
                 eta_ts_to_prop, eta_ts_to_charge, eta_therm, 
                 batteryPowerThroughput, charge_status, n):
    ''' Function that only runs when TS is running. Compute the power required'''
    
    new_eta_ts_to_prop = eta_ts_to_prop / eta_therm
    
    if n == 1 or 2 or 11 or 12 or 21 or 22:
        Preq = PowerToMaintainFlight / new_eta_ts_to_prop
        powerArray.append(Preq)
        
    elif n == 13 or 23 and charge_status is True:
        Preq = (PowerToMaintainFlight / new_eta_ts_to_prop) + (batteryPowerThroughput / eta_ts_to_charge)
        powerArray.append(Preq)
        
    elif n == 3 or 14: 
        Preq = PowerToMaintainFlight * ((1 - hybridizationFactor) / new_eta_ts_to_prop)
        powerArray.append(Preq)
        
    return powerArray
        
def fuel_burn(T, t, eta_ts_to_prop, eta_ts_to_charge, eta_batt_to_prop, eta_therm, 
              heatingValue, takeoffVelocity, hybridizationFactor,
              currentVelocity, batteryEmptyChargeCap, batteryFullChargeCap, 
              batteryPowerThroughput, missionRange, distanceCovered,
              t_idx, n, cap, charge_status, power_log):
    """
    Calculates the fuel burned based on the flight phase.
    Quick Guide:
        
        TAKEOFF
        case 1 = Turboshaft ONLY
        case 2 = Battery ONLY (if capable)
        case 3 = Quasi-Parallel
        
        CLIMB
        case 11 = Turboshaft ONLY
        case 12 = Battery till depleted, NO RECHARGE
        case 13 = Battery for Climb, RECHARGE
        case 14 = Quasi-parallel
        
        CRUISE
        case 21 = Battery at beginning of cruise, then turboshaft
        case 22 = Battery at end of mission
        case 23 = Cyclic Charging
    """
    
    match n:
        
        case 0:
            ''' Case to take care of time = 0 '''
            cap = cap
            m_fuel = 0
            P = 0
        
        
        
            ''' TAKEOFF OPTIONS '''
        case 1:                 
            # Takeoff turboshaft-only prop
            T_TO = 130290                               # thrust at takeoff [N]
            P = T_TO * takeoffVelocity               # power required for takeoff [W]
            
            m_fuel = (P * t) / (eta_ts_to_prop * heatingValue)
            cap = cap
            
            power_log = power_logger(power_log, P, hybridizationFactor, eta_ts_to_prop, 
                                     eta_ts_to_charge, eta_therm, 
                                     batteryPowerThroughput, 
                                     charge_status, n)
            
        case 2:
            # Takeoff battery primary prop
            T_TO = 130290                       # thrust at takeoff [N]
            P = T_TO * takeoffVelocity                  # power required for takeoff [W]
            
            # if sufficient capacity, then use battery only
            if cap >= P * 300: 
                m_fuel = 0
                cap -= (P * t) / eta_batt_to_prop
            
            # if no battery is present, use turboshaft only
            elif cap == 0:
                m_fuel = (P * t) / (eta_ts_to_prop * heatingValue)
                cap = 0
                
                power_log = power_logger(power_log, P, hybridizationFactor, eta_ts_to_prop, 
                                         eta_ts_to_charge, eta_therm, 
                                         batteryPowerThroughput, 
                                         charge_status, n)
                
            # if battery capacity is less than energy required for takeoff, use turboshaft 
            else:
                # Only turboshaft contributing to takeoff power if battery does not have enough capacity
                m_fuel = (P * t) / (eta_ts_to_prop * heatingValue)
                cap = 0
                
                power_log = power_logger(power_log, P, hybridizationFactor, eta_ts_to_prop, 
                                         eta_ts_to_charge, eta_therm, 
                                         batteryPowerThroughput, 
                                         charge_status, n)
                
        case 3:
            # Quasi-Parallel
            T_TO = 130290                       # thrust at takeoff [N]
            P = T_TO * takeoffVelocity             # power required for takeoff [W]
            
            # Quasi-Parallel Takeoff
            m_fuel = (P * (1 - hybridizationFactor) * t) / (eta_ts_to_prop * heatingValue)
            cap -= (P * t * hybridizationFactor) / eta_batt_to_prop
            
            power_log = power_logger(power_log, P, hybridizationFactor, eta_ts_to_prop, 
                                     eta_ts_to_charge, eta_therm, 
                                     batteryPowerThroughput, 
                                     charge_status, n)
            
            
            ''' CLIMB OPTIONS '''
        case 11: 
            # Climb turboshaft only
            P = T * currentVelocity  # power required to climb

            m_fuel = (P * t) / (eta_ts_to_prop * heatingValue)
            cap = cap
            
            power_log = power_logger(power_log, P, hybridizationFactor, eta_ts_to_prop, 
                                     eta_ts_to_charge, eta_therm, 
                                     batteryPowerThroughput, 
                                     charge_status, n)
            
        case 12:
            # Climb Battery then Turboshaft, no charging
            P = T * currentVelocity  # power required to climb
            
            # Battery first, then TS for remainder of climb
            if cap <= batteryEmptyChargeCap and cap != 0 or cap < \
            ((P * t) / eta_batt_to_prop):  # If battery is less than or equal to 20% charged and the capacity is non zero or the capacity is not enough to provide required energy
                
                m_fuel = (P * t) / (eta_ts_to_prop * heatingValue) 
                cap = cap
                charge_status = True
                power_log = power_logger(power_log, P, hybridizationFactor, eta_ts_to_prop, 
                                         eta_ts_to_charge, eta_therm, 
                                         batteryPowerThroughput, 
                                         charge_status, n)
                
            elif cap >= batteryEmptyChargeCap and cap != 0 and cap >= \
            ((P * t) / eta_batt_to_prop): # if battery has sufficient charge
                
                m_fuel = 0
                cap -= (P * t) / eta_batt_to_prop
            
            elif cap == 0:
                m_fuel = (P * t) / (eta_ts_to_prop * heatingValue)
                
                power_log = power_logger(power_log, P, hybridizationFactor, eta_ts_to_prop, 
                                         eta_ts_to_charge, eta_therm, 
                                         batteryPowerThroughput, 
                                         charge_status, n)
        case 13:
            # Climb Battery with recharging
            P = T * currentVelocity  # power required to climb
            
            # Charging Phase
            if cap <= batteryEmptyChargeCap or (cap - ((P * t) / eta_batt_to_prop)) <= 0:
                
                cap, m_fuel = charge_batt(P, t, 
                                          eta_ts_to_prop, 
                                          heatingValue, 
                                          batteryPowerThroughput, 
                                          eta_ts_to_charge, cap)
                charge_status = True                                        # To ensure that fuel is only used when charging
                
                power_log = power_logger(power_log, P, hybridizationFactor, eta_ts_to_prop, 
                                         eta_ts_to_charge, eta_therm, 
                                         batteryPowerThroughput, 
                                         charge_status, n)
                
            # Handle case where cap is between 30% and 100% but not at full capacity
            elif batteryEmptyChargeCap < cap < batteryFullChargeCap and charge_status is True:
                
                cap, m_fuel = charge_batt(P, t, 
                                          eta_ts_to_prop, 
                                          heatingValue, 
                                          batteryPowerThroughput, 
                                          eta_ts_to_charge, cap)
                       
                power_log = power_logger(power_log, P, hybridizationFactor, eta_ts_to_prop, 
                                         eta_ts_to_charge, eta_therm, 
                                         batteryPowerThroughput, 
                                         charge_status, n)
                        
                # Ensure the battery does not exceed full capacity
                if cap > batteryFullChargeCap:
                    m_fuel = m_fuel - (((cap - batteryFullChargeCap)) / (eta_ts_to_charge * heatingValue))
                    cap = batteryFullChargeCap  # Set to max capacity if exceeded
                            
            # Depleting Phase
            elif cap == batteryFullChargeCap:
                # No fuel burned while battery is being depleted
                m_fuel = 0
                
                # Deplete the battery until it reaches 30%
                cap -= (P * t) / eta_batt_to_prop # Deplete the battery
                
                charge_status = False
                    
                # Deplete battery without using fuel
            elif batteryEmptyChargeCap < cap < batteryFullChargeCap and charge_status is False:
                    m_fuel = 0;
                    
                    # Deplete the battery until it reaches 30%
                    cap -= (P * t) / eta_batt_to_prop  # Deplete the battery
                    
                    # Ensure the capacity does not fall below 30%
                    if cap < batteryEmptyChargeCap:
                        m_fuel = m_fuel + ((P * t) / (eta_ts_to_prop * heatingValue))
                        cap = batteryEmptyChargeCap # Set to minimum capacity if below  

                
        case 14:
            # Battery Supports TS in Climb, no recharging Option
            P = T * currentVelocity  # power required to climb
            
            if cap >= (P * t * hybridizationFactor) and cap > batteryEmptyChargeCap: 
                m_fuel = (P * (1 - hybridizationFactor) * t) / \
                    (eta_ts_to_prop * heatingValue)
                cap -= (P * t * hybridizationFactor) / eta_batt_to_prop
                
                power_log = power_logger(power_log, P, hybridizationFactor, eta_ts_to_prop, 
                                         eta_ts_to_charge, eta_therm, 
                                         batteryPowerThroughput, 
                                         charge_status, n)
            
            elif cap < (P * t * hybridizationFactor) or cap <= batteryEmptyChargeCap:
                m_fuel =  m_fuel = (P * t) / (eta_ts_to_prop * heatingValue)
                
                power_log = power_logger(power_log, P, hybridizationFactor, eta_ts_to_prop, 
                                         eta_ts_to_charge, eta_therm, 
                                         batteryPowerThroughput, 
                                         charge_status, n)
                
                # Ensure the capacity does not fall below 30%
                if cap < batteryEmptyChargeCap:
                    m_fuel = m_fuel + ((P * t) / (eta_ts_to_prop * heatingValue))
                    cap = batteryEmptyChargeCap # Set to minimum capacity if below  
                else: 
                    cap = cap
                charge_status = True
            
            
                
            ''' CRUISE OPTIONS '''
        case 21:
            # Battery used at beginning of cruise, then turboshaft the rest of the way
            P = T * currentVelocity  # power required for cruise
            
            if cap == batteryFullChargeCap:
                # No fuel burned while battery is being depleted
                m_fuel = 0
                
                # Deplete the battery until it reaches 30%
                cap -= (P * t) / eta_batt_to_prop  # Deplete the battery
                    
                # Deplete battery without using fuel
            elif batteryEmptyChargeCap < cap < batteryFullChargeCap:
                    m_fuel = 0;
                    
                    # Deplete the battery until it reaches 30%
                    cap -= (P * t) / eta_batt_to_prop # Deplete the battery
                    
                    # Ensure the capacity does not fall below 30%
                    if cap < batteryEmptyChargeCap:
                        cap = batteryEmptyChargeCap # Set to minimum capacity if below
                        m_fuel = m_fuel + ((P * t) / (eta_ts_to_prop * heatingValue))
                        
            elif cap <= batteryEmptyChargeCap or (cap - (P * t)) <= 0:
                m_fuel = (P * t) / (eta_ts_to_prop * heatingValue)
                
                power_log = power_logger(power_log, P, hybridizationFactor, eta_ts_to_prop, 
                                         eta_ts_to_charge, eta_therm, 
                                         batteryPowerThroughput, 
                                         charge_status, n)
                
                # Ensure the capacity does not fall below 30%
                if cap < batteryEmptyChargeCap:
                    cap = batteryEmptyChargeCap # Set to minimum capacity if below
                    m_fuel = m_fuel + ((P * t) / (eta_ts_to_prop * heatingValue))
                
            
        case 22:
            # Turboshaft first, then battery rest of the way
            P = T * currentVelocity  # power required for cruise
            
            timeRemaining = (missionRange - distanceCovered) / currentVelocity
            
            energyReqRemaining = (P * timeRemaining) / eta_batt_to_prop 
            
            if cap > energyReqRemaining and cap > batteryEmptyChargeCap:
                m_fuel = 0
                cap -= (P * t) / eta_batt_to_prop
                
            else:
                m_fuel = (P * t) / (eta_ts_to_prop * heatingValue)
                cap = cap
                
                power_log = power_logger(power_log, P, hybridizationFactor, eta_ts_to_prop, 
                                         eta_ts_to_charge, eta_therm, 
                                         batteryPowerThroughput, 
                                         charge_status, n)
           
        case 23:
            # Cyclic Charging
            P = T * currentVelocity  # power required for cruise
            
            t_batt_afterClimbTO = cap / (P * eta_batt_to_prop)
            
            t_elapsed_after_climb = t + t_idx
            
            if t_elapsed_after_climb < t_batt_afterClimbTO:                     # No fuel burned if battery is still usable
                m_fuel = 0
                cap -= (P * t) / eta_batt_to_prop
            elif cap == 0 or cap < (P * t):
                m_fuel = (P * t) / (eta_ts_to_prop * heatingValue)
                
                power_log = power_logger(power_log, P, hybridizationFactor, eta_ts_to_prop, 
                                         eta_ts_to_charge, eta_therm, 
                                         batteryPowerThroughput, 
                                         charge_status, n)
                cap = cap                                   
            else:
                # Charging Phase
                 if cap <= batteryEmptyChargeCap or (cap - ((P * t) / eta_batt_to_prop)) <= 0:
                    
                    cap, m_fuel = charge_batt(P, t, 
                                              eta_ts_to_prop, 
                                              heatingValue, 
                                              batteryPowerThroughput, 
                                              eta_ts_to_charge, cap)
                    
                    charge_status = True                                        # To ensure that fuel is only used when charging
                    
                    power_log = power_logger(power_log, P, hybridizationFactor, eta_ts_to_prop, 
                                             eta_ts_to_charge, eta_therm, 
                                             batteryPowerThroughput, 
                                             charge_status, n)
                    
                # Handle case where cap is between 30% and 100% but not at full capacity
                 elif batteryEmptyChargeCap < cap < batteryFullChargeCap and charge_status is True:
                    
                    cap, m_fuel = charge_batt(P, t, 
                                              eta_ts_to_prop, 
                                              heatingValue, 
                                              batteryPowerThroughput, 
                                              eta_ts_to_charge, cap)
                    
                    power_log = power_logger(power_log, P, hybridizationFactor, eta_ts_to_prop, 
                                             eta_ts_to_charge, eta_therm, 
                                             batteryPowerThroughput, 
                                             charge_status, n)
                        
                    # Ensure the battery does not exceed full capacity
                    if cap > batteryFullChargeCap:
                        m_fuel = m_fuel - (((cap - batteryFullChargeCap)) / (eta_ts_to_charge * heatingValue))
                        cap = batteryFullChargeCap  # Set to max capacity if exceeded
                            
                # Depleting Phase
                 elif cap == batteryFullChargeCap:
                    # No fuel burned while battery is being depleted
                    m_fuel = 0
                    
                    # Deplete the battery until it reaches 30%
                    cap -= (P * t) / eta_batt_to_prop # Deplete the battery
                    
                    charge_status = False
                        
                    # Deplete battery without using fuel
                 elif batteryEmptyChargeCap < cap < batteryFullChargeCap and charge_status is False:
                        m_fuel = 0;
                        
                        # Deplete the battery until it reaches 30%
                        cap -= (P * t) / eta_batt_to_prop  # Deplete the battery
                        
                        # Ensure the capacity does not fall below 30%
                        if cap < batteryEmptyChargeCap:
                            m_fuel = m_fuel + ((P * t) / (eta_ts_to_prop * heatingValue))
                            cap = batteryEmptyChargeCap # Set to minimum capacity if below  
     

    return m_fuel, cap, charge_status, power_log, P

def drag_polar_parabolic(L, q, S, c_d0, pi_AR_e):
    """
    Calculate the drag polar to obtain lift-to-drag ratio.
    """
    c_l = L / (q * S) # Coefficient of lift
    if c_l > 100:
        print(f"Warning: c_l value is unusually large: {c_l}")
    c_d = c_d0 + (c_l ** 2) / (pi_AR_e) # Coefficient of drag
    L_D = c_l / c_d # Lift-to-drag ratio
    return L_D

def drag_polar_table(L, q, S):
    path = os.path.join('..', 'input_data', 'drag_polar_tab.csv')
    
    if not os.path.exists(path):
        raise FileNotFoundError(f"The file '{path}' does not exist.")
        
    c_l = L / (q * S)  # Coefficient of lift
    c_l_rounded = round(c_l, 3)
    dragPolarTable = pd.read_csv(path)

    # Try to find the matching c_d
    result = dragPolarTable[dragPolarTable['c_l'] == c_l_rounded]['c_d']

    # Check if result is empty (no match found)
    if result.size == 0:
        raise ValueError(f"No Drag Coefficient found for c_l = {c_l_rounded}.")
    else:
        c_d = result.values[0]  # Get the first matching value
        L_D = c_l/c_d
        
    return L_D


def add_L_D_to_3d_deque(L_D_3d_deque, slice_idx, col_idx, L_D_value):
    """
    Adds an L/D value to the 3D deque at the specified slice (slice_idx) and column (col_idx).
    The number of rows (innermost deque) is dynamic.
    """
    # Append the L/D value to the specified column's row deque
    L_D_3d_deque[slice_idx][col_idx].append(L_D_value)
    

def fuel_check(M_fuel_init, m_fuel):
    """ Check if fuel burned exceeds the initial amount of fuel onboard aircraft"""
    if m_fuel > M_fuel_init:
        print("FUEL BURNED EXCEEDS CAPACITY!")
        
def initial_fuel_calculator(AC, payloadMass=None, batt_m=None, payloadMassArray=None, batteryMassArray=None, specificEnergyArray=None, specificEnergy=None):
    
    if payloadMassArray is None:
        if batteryMassArray is not None and specificEnergyArray is not None:
            FuelMass = np.zeros(len(batteryMassArray))
            for u, batt_mass in enumerate(batteryMassArray):
                for specificEnergy in specificEnergyArray:
                    if specificEnergy == 0:
                        FuelMass[u] = AC["MTOW"] - payloadMass - AC["OperatingEmptyMass"]
                    else:
                        FuelMass[u] = AC["MTOW"] - payloadMass - AC["OperatingEmptyMass"] - batt_mass
            fuel_to_batt_ratio = FuelMass / batteryMassArray
            
        elif batteryMassArray is not None:
            FuelMass = np.zeros(len(batteryMassArray))
            for z, batt_mass in enumerate(batteryMassArray):
                FuelMass[z] = AC["MTOW"] - payloadMass - AC["OperatingEmptyMass"] - (0 if specificEnergy == 0 else batt_mass)
            fuel_to_batt_ratio = FuelMass / batteryMassArray
            
        elif specificEnergyArray is not None:
            FuelMass = np.zeros([2])
            u = 0
            FuelMass[0] = AC["MTOW"] - payloadMass - AC["OperatingEmptyMass"]
            FuelMass[1] = AC["MTOW"] - payloadMass - AC["OperatingEmptyMass"] - batt_m
            fuel_to_batt_ratio = FuelMass / batt_m
            
        else:
            FuelMass = AC["MTOW"] - payloadMass - AC["OperatingEmptyMass"] - (0 if specificEnergy == 0 else batt_m)
            fuel_to_batt_ratio = FuelMass / batt_m
    else:
        if batteryMassArray is not None and specificEnergyArray is not None:
            FuelMass = np.zeros((len(batteryMassArray), len(payloadMassArray)))
            for u, batt_mass in enumerate(batteryMassArray):
                for specificEnergy in specificEnergyArray:
                    for z, payloadMass in enumerate(payloadMassArray):
                        if specificEnergy == 0:
                            FuelMass[u, z] = AC["MTOW"] - payloadMass - AC["OperatingEmptyMass"]
                        else:
                            FuelMass[u, z] = AC["MTOW"] - payloadMass - AC["OperatingEmptyMass"] - batt_mass
            fuel_to_batt_ratio = FuelMass / batteryMassArray
            
        elif batteryMassArray is not None:
            FuelMass = np.zeros((len(batteryMassArray), len(payloadMassArray)))
            for u, batt_mass in enumerate(batteryMassArray):
                for z, payloadMass in enumerate(payloadMassArray):
                    FuelMass[u, z] = AC["MTOW"] - payloadMass - AC["OperatingEmptyMass"] - (0 if specificEnergy == 0 else batt_mass)
            fuel_to_batt_ratio = FuelMass / batteryMassArray
            
        elif specificEnergyArray is not None:
            FuelMass = np.zeros([2, len(payloadMassArray)])
            for z, payloadMass in enumerate(payloadMassArray):
                FuelMass[0, z] = AC["MTOW"] - payloadMass - AC["OperatingEmptyMass"]
                FuelMass[1, z] = AC["MTOW"] - payloadMass - AC["OperatingEmptyMass"] - batt_m
            fuel_to_batt_ratio = FuelMass / batt_m
            
        else:
            FuelMass = np.zeros([len(payloadMassArray)])
            for z, payloadMass in enumerate(payloadMassArray):
                FuelMass[z] = AC["MTOW"] - payloadMass - AC["OperatingEmptyMass"] - (0 if specificEnergy == 0 else batt_m)
            fuel_to_batt_ratio = FuelMass / batt_m
            
    return FuelMass, fuel_to_batt_ratio
        
class CarpetPlot:
    def __init__(self, original_data):
        self.original_data = original_data
        self.modified_data = None
        self.staggered_data = None

    def create_modified_data(self):
        """ Modify the original matrix to include n extra rows and columns. """
        p = 0
        rows, cols = self.original_data.shape
        if rows != cols:
            sys.exit('ERROR: Matrix is asymmetric!')
            
        else:
            self.modified_data = np.zeros((rows, cols + cols), dtype=self.original_data.dtype)
            for j in range((cols+cols)):
                for i in range(rows):
                    if j < cols:
                        self.modified_data[i, j] = self.original_data[i, j]             # Original data remains in the first part
                    
                    # Assign values for the second part based on the specified pattern
                    if j >= cols:  # Ensure we don't go out of bounds
                        self.modified_data[i, j] = self.original_data[p,((cols-1)-i)]  # Get next row's value
                if j >= cols:
                    p += 1

    def create_staggered_data(self):
        """ Stagger the modified matrix and prepare it for Excel. """
        if self.modified_data is None:
            raise ValueError("Modified matrix not created. Call create_modified_data first.")

        rows, cols = self.modified_data.shape
        staggered_rows = rows + (rows - 1)
        # Create a new array with an extra column for the row numbers
        self.staggered_data = np.zeros((staggered_rows, cols + 1), dtype=self.modified_data.dtype)
 
        # Fill the first column with row numbers
        for i in range(staggered_rows):
            self.staggered_data[i, 0] = i + 1  # Row numbers starting from 1
            
        u = 1
        for j in range(cols):
            for i in range(rows):
                if j < (rows-1):
                    self.staggered_data[i + (rows-(1+j)), j+1] = self.modified_data[i, j]
                elif j == rows or j == (rows-1):
                    self.staggered_data[i, j+1] = self.modified_data[i, j]
                elif j > rows:
                    self.staggered_data[i+u, j+1] = self.modified_data[i, j]
            if j > rows:
                u += 1
            
    def to_excel(self, filename, sheet_name):
        """ Export the staggered matrix to an existing Excel file or create a new one. """
        if self.staggered_data is None:
            raise ValueError("Staggered matrix not created. Call create_staggered_matrix first.")
    
        staggered_df = pd.DataFrame(self.staggered_data).replace(0, np.nan)
    
        filepath = os.path.join('..', 'output_data', 'plots', filename)
        
        try:
            # Load the existing workbook
            book = load_workbook(filepath)
            
            # Check if all sheets are hidden
            all_sheets_hidden = all(book[sheet].sheet_state == 'hidden' for sheet in book.sheetnames)
            
            if all_sheets_hidden:
                # Unhide the first sheet if all are hidden
                first_sheet = book.active  # Get the active sheet
                first_sheet.sheet_state = 'visible'  # Unhide it
            
            # Check if the sheet already exists
            if sheet_name in book.sheetnames:
                raise ValueError(f"Sheet '{sheet_name}' already exists. Please choose a different name.")
            
            # Write to a new sheet by appending to the existing file
            with pd.ExcelWriter(filepath, engine='openpyxl', mode='a') as writer:
                staggered_df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)
        
        except FileNotFoundError:
            # If the file does not exist, create a new one
            staggered_df.to_excel(filepath, sheet_name=sheet_name, index=False, header=False)
    
        print(f"Staggered matrix exported to {filepath} in sheet '{sheet_name}'")
